"""
Account app: facade over core (org). Org-level overview, organization, billing summary, connectors.
Executive-only for Account; Manager manages billing. See docs/Account app - design and feedback.md.
"""
import uuid as uuid_module
from django.utils.crypto import get_random_string
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

from users.permission_classes import HasFeaturePermission


def _get_organization_id(request):
    """Resolve current user's organization_id from UserProfile."""
    if not request.user.is_authenticated:
        return None
    profile = getattr(request.user, "profile", None)
    if profile is None:
        return None
    return getattr(profile, "organization_id", None)


def _connector_to_json(connector, modules_list=None):
    """Serialize OrgConnector to API shape; modules_list from prefetched module_scopes or query."""
    from platform_org.models import OrgConnectorModuleScope
    if modules_list is None:
        modules_list = list(
            OrgConnectorModuleScope.objects.using("org")
            .filter(org_connector_id=connector.id)
            .values_list("module", flat=True)
        )
    return {
        "id": str(connector.id),
        "name": connector.name,
        "connector_type": connector.connector_type,
        "governance_set": connector.governance_set,
        "modules": modules_list,
        "config": connector.config or {},
        "credentials_ref": connector.credentials_ref or "",
        "status": connector.status or "",
        "last_synced_at": connector.last_synced_at.isoformat() if connector.last_synced_at else None,
        "created_at": connector.created_at.isoformat() if connector.created_at else None,
        "updated_at": connector.updated_at.isoformat() if connector.updated_at else None,
    }


@api_view(["GET"])
@permission_classes([IsAuthenticated, HasFeaturePermission("account.overview.view")])
def overview(request):
    """
    Account overview: org summary, billing summary, security summary.
    Requires account.overview.view (Executive).
    """
    org_id = _get_organization_id(request)
    data = {"organization": None, "billing_summary": None, "security_summary": None}

    if org_id:
        try:
            from platform_org.models import Organization, Team, Environment
            org = Organization.objects.using("org").filter(id=org_id).first()
            if org:
                data["organization"] = {
                    "id": str(org.id),
                    "org_id": org.org_id,
                    "name": org.name,
                    "tier": org.tier,
                }
            teams = []
            envs = []
            if org:
                teams = list(Team.objects.using("org").filter(organization_id=org.id).values_list("id", flat=True))
                envs = list(Environment.objects.using("org").filter(organization_id=org.id).values_list("id", flat=True))
            data["organization_structure"] = {
                "teams_count": len(teams),
                "environments_count": len(envs),
            }
        except Exception:
            pass

    # Billing summary (current user's subscription from financials)
    try:
        from financials.models import UserSubscription
        sub = UserSubscription.objects.filter(user=request.user, status="active").order_by("-created_at").first()
        if sub:
            data["billing_summary"] = {
                "plan_name": sub.plan_name,
                "status": sub.status,
                "billing_period": sub.billing_period,
                "next_billing_date": sub.next_billing_date.isoformat() if sub.next_billing_date else None,
            }
    except Exception:
        pass

    # Security summary (from UserProfile)
    try:
        profile = getattr(request.user, "profile", None)
        if profile:
            data["security_summary"] = {
                "two_factor_enabled": getattr(profile, "two_factor_enabled", False),
                "last_login": profile.last_login.isoformat() if getattr(profile, "last_login", None) else None,
            }
    except Exception:
        pass

    return Response(data)


@api_view(["GET", "PATCH"])
@permission_classes([IsAuthenticated, HasFeaturePermission("account.overview.view")])
def organization(request):
    """
    Get or update current organization. Requires account.overview.view (Executive).
    """
    org_id = _get_organization_id(request)
    if not org_id:
        return Response(
            {"error": "No organization assigned. Set profile.organization_id."},
            status=status.HTTP_404_NOT_FOUND,
        )

    from platform_org.models import Organization
    org = Organization.objects.using("org").filter(id=org_id).first()
    if not org:
        return Response({"error": "Organization not found."}, status=status.HTTP_404_NOT_FOUND)

    def _org_to_dict(o):
        return {
            "id": str(o.id),
            "org_id": o.org_id,
            "name": o.name,
            "tier": o.tier,
            "legal_name": o.legal_name,
            "display_name": o.display_name,
            "primary_domain": o.primary_domain,
            "secondary_domains": o.secondary_domains,
            "physical_address": o.physical_address,
            "industry": o.industry,
            "jurisdiction_country": o.jurisdiction_country,
            "jurisdiction_state": o.jurisdiction_state,
            "timezone": o.timezone,
            "currency": o.currency,
            "fiscal_year_start_month": o.fiscal_year_start_month,
            "primary_contact_name": o.primary_contact_name,
            "primary_contact_email": o.primary_contact_email,
            "billing_contact_email": o.billing_contact_email,
            "security_contact_email": o.security_contact_email,
            "data_residency_region": o.data_residency_region,
            "default_environment_id": str(o.default_environment_id) if o.default_environment_id else None,
            "enforce_mfa": o.enforce_mfa,
            "sso_enabled": o.sso_enabled,
            "status": o.status,
            "created_at": o.created_at.isoformat() if o.created_at else None,
            "created_by": str(o.created_by) if o.created_by else None,
            "changed_by": str(o.changed_by) if o.changed_by else None,
            "changed_at": o.changed_at.isoformat() if o.changed_at else None,
        }

    if request.method == "GET":
        return Response(_org_to_dict(org))

    if request.method == "PATCH":
        from django.utils import timezone as tz
        TEXT_FIELDS = [
            "name", "tier", "legal_name", "display_name", "primary_domain",
            "secondary_domains", "physical_address", "industry",
            "jurisdiction_country", "jurisdiction_state", "timezone", "currency",
            "primary_contact_name", "primary_contact_email",
            "billing_contact_email", "security_contact_email",
            "data_residency_region", "status",
        ]
        for field in TEXT_FIELDS:
            val = request.data.get(field)
            if val is not None:
                setattr(org, field, val)

        fys = request.data.get("fiscal_year_start_month")
        if fys is not None:
            org.fiscal_year_start_month = int(fys)

        for bool_field in ("enforce_mfa", "sso_enabled"):
            val = request.data.get(bool_field)
            if val is not None:
                setattr(org, bool_field, bool(val))

        env_id = request.data.get("default_environment_id")
        if env_id is not None:
            org.default_environment_id = env_id or None

        org.changed_by = request.user.id if hasattr(request.user, "id") else None
        org.changed_at = tz.now()
        org.save(using="org")
        return Response(_org_to_dict(org))


@api_view(["GET"])
@permission_classes([IsAuthenticated, HasFeaturePermission("account.billing.view")])
def billing_summary(request):
    """
    Billing summary for Account (org context). Manager and Executive can view.
    Returns current user's active subscription; later can scope by org.
    """
    try:
        from financials.models import UserSubscription
        subs = UserSubscription.objects.filter(user=request.user).order_by("-created_at")[:5]
        data = {
            "subscriptions": [
                {
                    "id": s.id,
                    "plan_name": s.plan_name,
                    "status": s.status,
                    "billing_period": s.billing_period,
                    "next_billing_date": s.next_billing_date.isoformat() if s.next_billing_date else None,
                }
                for s in subs
            ],
        }
        return Response(data)
    except Exception as e:
        return Response({"subscriptions": [], "error": str(e)}, status=status.HTTP_200_OK)


# --- Connectors (Set 2 & 3): org-scoped CRUD; only Account can write ---

@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated, HasFeaturePermission("account.overview.view")])
def connector_list(request):
    """
    List connectors for current org (GET) or create one (POST).
    Requires account.overview.view.
    """
    org_id = _get_organization_id(request)
    if not org_id:
        return Response(
            {"error": "No organization assigned."},
            status=status.HTTP_404_NOT_FOUND,
        )

    from platform_org.models import OrgConnector, OrgConnectorModuleScope

    if request.method == "GET":
        connectors = (
            OrgConnector.objects.using("org")
            .filter(organization_id=org_id)
            .order_by("connector_type", "name")
        )
        connector_ids = [c.id for c in connectors]
        scope_map = {}
        for s in OrgConnectorModuleScope.objects.using("org").filter(org_connector_id__in=connector_ids).values("org_connector_id", "module"):
            scope_map.setdefault(str(s["org_connector_id"]), []).append(s["module"])
        out = []
        for c in connectors:
            out.append(_connector_to_json(c, modules_list=scope_map.get(str(c.id), [])))
        return Response(out)

    if request.method == "POST":
        name = (request.data.get("name") or "").strip()
        connector_type = request.data.get("connector_type")
        governance_set = request.data.get("governance_set", 3)
        modules = request.data.get("modules") or []
        config = request.data.get("config") or {}
        credentials_ref = (request.data.get("credentials_ref") or "").strip()
        if not name:
            return Response({"error": "name is required."}, status=status.HTTP_400_BAD_REQUEST)
        if connector_type not in dict(OrgConnector.CONNECTOR_TYPE_CHOICES):
            return Response({"error": "Invalid connector_type."}, status=status.HTTP_400_BAD_REQUEST)
        if governance_set not in (2, 3):
            return Response({"error": "governance_set must be 2 or 3."}, status=status.HTTP_400_BAD_REQUEST)
        connector = OrgConnector(
            organization_id=org_id,
            name=name,
            connector_type=connector_type,
            governance_set=governance_set,
            config=config,
            credentials_ref=credentials_ref,
            status="active",
        )
        connector.save(using="core")
        valid_modules = [m[0] for m in OrgConnectorModuleScope.MODULE_CHOICES]
        for mod in modules:
            if mod in valid_modules:
                OrgConnectorModuleScope.objects.using("org").create(
                    org_connector_id=connector.id,
                    module=mod,
                )
        return Response(_connector_to_json(connector), status=status.HTTP_201_CREATED)


@api_view(["GET", "PATCH", "DELETE"])
@permission_classes([IsAuthenticated, HasFeaturePermission("account.overview.view")])
def connector_detail(request, connector_id):
    """
    Get (GET), update (PATCH), or delete (DELETE) a connector. Org-scoped.
    """
    org_id = _get_organization_id(request)
    if not org_id:
        return Response(
            {"error": "No organization assigned."},
            status=status.HTTP_404_NOT_FOUND,
        )

    from platform_org.models import OrgConnector, OrgConnectorModuleScope

    connector = (
        OrgConnector.objects.using("org")
        .filter(id=connector_id, organization_id=org_id)
        .first()
    )
    if not connector:
        return Response({"error": "Connector not found."}, status=status.HTTP_404_NOT_FOUND)

    if request.method == "GET":
        modules_list = list(
            OrgConnectorModuleScope.objects.using("org")
            .filter(org_connector_id=connector.id)
            .values_list("module", flat=True)
        )
        return Response(_connector_to_json(connector, modules_list=modules_list))

    if request.method == "PATCH":
        name = request.data.get("name")
        if name is not None:
            connector.name = (name or "").strip()
            if not connector.name:
                return Response({"error": "name cannot be empty."}, status=status.HTTP_400_BAD_REQUEST)
        if "governance_set" in request.data:
            gs = request.data["governance_set"]
            if gs not in (2, 3):
                return Response({"error": "governance_set must be 2 or 3."}, status=status.HTTP_400_BAD_REQUEST)
            connector.governance_set = gs
        if "config" in request.data:
            connector.config = request.data["config"] or {}
        if "credentials_ref" in request.data:
            connector.credentials_ref = (request.data["credentials_ref"] or "").strip()
        if "status" in request.data:
            connector.status = (request.data["status"] or "").strip()
        connector.save(using="core")
        modules = request.data.get("modules")
        if modules is not None:
            OrgConnectorModuleScope.objects.using("org").filter(org_connector_id=connector.id).delete()
            valid_modules = [m[0] for m in OrgConnectorModuleScope.MODULE_CHOICES]
            for mod in modules:
                if mod in valid_modules:
                    OrgConnectorModuleScope.objects.using("org").create(
                        org_connector_id=connector.id,
                        module=mod,
                    )
        modules_list = list(
            OrgConnectorModuleScope.objects.using("org")
            .filter(org_connector_id=connector.id)
            .values_list("module", flat=True)
        )
        return Response(_connector_to_json(connector, modules_list=modules_list))

    if request.method == "DELETE":
        connector.delete(using="core")
        return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def connector_scope(request):
    """
    List connectors visible to a module (scoping logic). Read-only.
    GET /api/account/connectors/scope/?module=discovery
    Returns connectors for current org where governance_set=2 (all modules)
    or governance_set=3 and this module is in scope.
    """
    org_id = _get_organization_id(request)
    if not org_id:
        return Response(
            {"error": "No organization assigned."},
            status=status.HTTP_404_NOT_FOUND,
        )
    module = (request.query_params.get("module") or "").strip().lower()
    if not module:
        return Response({"error": "Query param 'module' is required (e.g. discovery, security)."}, status=status.HTTP_400_BAD_REQUEST)

    from platform_org.models import OrgConnector, OrgConnectorModuleScope

    valid_modules = [m[0] for m in OrgConnectorModuleScope.MODULE_CHOICES]
    if module not in valid_modules:
        return Response({"error": f"Invalid module. Must be one of: {', '.join(valid_modules)}."}, status=status.HTTP_400_BAD_REQUEST)

    # Set 2: all connectors with governance_set=2
    global_ids = set(
        OrgConnector.objects.using("org")
        .filter(organization_id=org_id, governance_set=2)
        .values_list("id", flat=True)
    )
    # Set 3: connectors that have this module in scope
    scoped_ids = set(
        OrgConnectorModuleScope.objects.using("org")
        .filter(org_connector__organization_id=org_id, org_connector__governance_set=3, module=module)
        .values_list("org_connector_id", flat=True)
    )
    all_ids = list(global_ids | scoped_ids)
    connectors = OrgConnector.objects.using("org").filter(id__in=all_ids).order_by("connector_type", "name")
    scope_map = {}
    for s in OrgConnectorModuleScope.objects.using("org").filter(org_connector_id__in=all_ids).values("org_connector_id", "module"):
        scope_map.setdefault(str(s["org_connector_id"]), []).append(s["module"])
    out = [_connector_to_json(c, modules_list=scope_map.get(str(c.id), [])) for c in connectors]
    return Response(out)


# --- Multi-region: countries (reference) + organization regions (per-org) ---

@api_view(["GET"])
@permission_classes([IsAuthenticated, HasFeaturePermission("account.overview.view")])
def country_list(request):
    """
    List countries for multi-region dropdowns. Reference data (EU, US, APAC, etc.).
    """
    from platform_org.models import Country
    countries = Country.objects.using("org").order_by("name").values("id", "code", "name", "region_group")
    return Response([{"id": str(c["id"]), "code": c["code"], "name": c["name"], "region_group": c["region_group"]} for c in countries])


@api_view(["GET"])
@permission_classes([IsAuthenticated, HasFeaturePermission("account.overview.view")])
def state_list(request):
    """List states/provinces for a country. Query param: ?country=US"""
    from platform_org.models import State
    country_code = request.query_params.get("country", "").strip().upper()
    if not country_code:
        return Response([])
    states = (
        State.objects.using("org")
        .filter(country__code=country_code)
        .order_by("name")
        .values("id", "code", "name")
    )
    return Response([{"id": str(s["id"]), "code": s["code"], "name": s["name"]} for s in states])


@api_view(["GET"])
@permission_classes([IsAuthenticated, HasFeaturePermission("account.overview.view")])
def currency_list(request):
    """List currencies for org settings dropdowns."""
    from platform_org.models import Currency
    currencies = Currency.objects.using("org").order_by("code").values("id", "code", "name", "symbol")
    return Response([{"id": str(c["id"]), "code": c["code"], "name": c["name"], "symbol": c["symbol"]} for c in currencies])


@api_view(["GET"])
@permission_classes([IsAuthenticated, HasFeaturePermission("account.overview.view")])
def timezone_list(request):
    """List timezones for org settings dropdowns."""
    from platform_org.models import Timezone
    tzs = Timezone.objects.using("org").order_by("utc_offset", "name").values("id", "name", "label", "utc_offset")
    return Response([{"id": str(t["id"]), "name": t["name"], "label": t["label"], "utc_offset": t["utc_offset"]} for t in tzs])


def _org_region_to_json(region):
    return {
        "id": str(region.id),
        "country_id": str(region.country_id),
        "country_code": region.country.code,
        "country_name": region.country.name,
        "region_group": region.country.region_group,
        "name": region.name,
        "city": region.city or "",
        "currency_code": region.currency_code or "",
        "primary_contact_name": region.primary_contact_name or "",
        "primary_contact_email": region.primary_contact_email or "",
        "address": region.address or "",
        "created_at": region.created_at.isoformat() if region.created_at else None,
        "updated_at": region.updated_at.isoformat() if region.updated_at else None,
    }


@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated, HasFeaturePermission("account.overview.view")])
def organization_region_list(request):
    """
    List (GET) or create (POST) organization regions for current org. Multi-region compliance (EU vs US).
    """
    org_id = _get_organization_id(request)
    if not org_id:
        return Response(
            {"error": "No organization assigned."},
            status=status.HTTP_404_NOT_FOUND,
        )

    from platform_org.models import OrganizationRegion, Country

    if request.method == "GET":
        regions = (
            OrganizationRegion.objects.using("org")
            .filter(organization_id=org_id)
            .order_by("name")
        )
        return Response([_org_region_to_json(r) for r in regions])

    if request.method == "POST":
        country_id = request.data.get("country_id")
        if not country_id:
            return Response({"error": "country_id is required."}, status=status.HTTP_400_BAD_REQUEST)
        country = Country.objects.using("org").filter(id=country_id).first()
        if not country:
            return Response({"error": "Invalid country_id."}, status=status.HTTP_400_BAD_REQUEST)
        if OrganizationRegion.objects.using("org").filter(organization_id=org_id, country_id=country_id).exists():
            return Response({"error": "A region for this country already exists."}, status=status.HTTP_400_BAD_REQUEST)
        city = (request.data.get("city") or "").strip()
        currency_code = (request.data.get("currency_code") or "").strip().upper()
        name = (request.data.get("name") or "").strip() or f"{country.name} operations"
        region = OrganizationRegion.objects.using("org").create(
            organization_id=org_id,
            country_id=country_id,
            name=name,
            city=city,
            currency_code=currency_code,
            primary_contact_name=(request.data.get("primary_contact_name") or "").strip(),
            primary_contact_email=(request.data.get("primary_contact_email") or "").strip(),
            address=(request.data.get("address") or "").strip(),
        )
        return Response(_org_region_to_json(region), status=status.HTTP_201_CREATED)


@api_view(["GET", "PATCH", "DELETE"])
@permission_classes([IsAuthenticated, HasFeaturePermission("account.overview.view")])
def organization_region_detail(request, region_id):
    """
    Get (GET), update (PATCH), or delete (DELETE) an organization region. Org-scoped.
    """
    org_id = _get_organization_id(request)
    if not org_id:
        return Response(
            {"error": "No organization assigned."},
            status=status.HTTP_404_NOT_FOUND,
        )

    from platform_org.models import OrganizationRegion

    region = (
        OrganizationRegion.objects.using("org")
        .filter(id=region_id, organization_id=org_id)
        .first()
    )
    if not region:
        return Response({"error": "Region not found."}, status=status.HTTP_404_NOT_FOUND)

    if request.method == "GET":
        return Response(_org_region_to_json(region))

    if request.method == "PATCH":
        if "name" in request.data:
            region.name = (request.data.get("name") or "").strip() or region.name
        if "city" in request.data:
            region.city = (request.data.get("city") or "").strip()
        if "currency_code" in request.data:
            region.currency_code = (request.data.get("currency_code") or "").strip().upper()
        if "primary_contact_name" in request.data:
            region.primary_contact_name = (request.data.get("primary_contact_name") or "").strip()
        if "primary_contact_email" in request.data:
            region.primary_contact_email = (request.data.get("primary_contact_email") or "").strip()
        if "address" in request.data:
            region.address = (request.data.get("address") or "").strip()
        region.save(using="org")
        return Response(_org_region_to_json(region))

    if request.method == "DELETE":
        region.delete(using="org")
        return Response(status=status.HTTP_204_NO_CONTENT)


# --- Multi-currency: org-enabled currencies ---

def _org_currency_to_json(oc):
    return {
        "id": str(oc.id),
        "currency_id": str(oc.currency_id),
        "code": oc.currency.code,
        "name": oc.currency.name,
        "symbol": oc.currency.symbol,
        "is_default": oc.is_default,
        "created_at": oc.created_at.isoformat() if oc.created_at else None,
    }


@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated, HasFeaturePermission("account.overview.view")])
def org_currency_list(request):
    """List (GET) or add (POST) currencies for current org."""
    org_id = _get_organization_id(request)
    if not org_id:
        return Response({"error": "No organization assigned."}, status=status.HTTP_404_NOT_FOUND)

    from platform_org.models import OrgCurrency, Currency

    if request.method == "GET":
        items = (
            OrgCurrency.objects.using("org")
            .filter(organization_id=org_id)
            .select_related("currency")
            .order_by("-is_default", "currency__code")
        )
        return Response([_org_currency_to_json(oc) for oc in items])

    if request.method == "POST":
        currency_id = request.data.get("currency_id")
        if not currency_id:
            return Response({"error": "currency_id is required."}, status=status.HTTP_400_BAD_REQUEST)
        cur = Currency.objects.using("org").filter(id=currency_id).first()
        if not cur:
            return Response({"error": "Invalid currency_id."}, status=status.HTTP_400_BAD_REQUEST)
        if OrgCurrency.objects.using("org").filter(organization_id=org_id, currency_id=currency_id).exists():
            return Response({"error": "Currency already added."}, status=status.HTTP_400_BAD_REQUEST)
        is_default = bool(request.data.get("is_default", False))
        if is_default:
            OrgCurrency.objects.using("org").filter(organization_id=org_id, is_default=True).update(is_default=False)
        oc = OrgCurrency.objects.using("org").create(
            organization_id=org_id,
            currency_id=currency_id,
            is_default=is_default,
        )
        oc.currency = cur
        return Response(_org_currency_to_json(oc), status=status.HTTP_201_CREATED)


@api_view(["PATCH", "DELETE"])
@permission_classes([IsAuthenticated, HasFeaturePermission("account.overview.view")])
def org_currency_detail(request, oc_id):
    """Update (PATCH — set default) or remove (DELETE) an org currency."""
    org_id = _get_organization_id(request)
    if not org_id:
        return Response({"error": "No organization assigned."}, status=status.HTTP_404_NOT_FOUND)

    from platform_org.models import OrgCurrency

    oc = OrgCurrency.objects.using("org").filter(id=oc_id, organization_id=org_id).select_related("currency").first()
    if not oc:
        return Response({"error": "Not found."}, status=status.HTTP_404_NOT_FOUND)

    if request.method == "PATCH":
        is_default = request.data.get("is_default")
        if is_default is not None and bool(is_default):
            OrgCurrency.objects.using("org").filter(organization_id=org_id, is_default=True).update(is_default=False)
            oc.is_default = True
            oc.save(using="org")
        return Response(_org_currency_to_json(oc))

    if request.method == "DELETE":
        oc.delete(using="org")
        return Response(status=status.HTTP_204_NO_CONTENT)


# --- Org-scoped users (Users & Access) ---

def _user_to_org_json(user):
    """Serialize user for org Users & Access list."""
    from users.permission_utils import _get_user_role_group
    profile = getattr(user, "profile", None)
    role = "viewer"
    if user.is_superuser:
        role = "admin"
    elif profile:
        group = _get_user_role_group(user)
        role = (group.name.lower() if group else profile.role) or "viewer"
    title = ""
    if profile and isinstance(profile.user_settings, dict):
        title = profile.user_settings.get("title", "") or ""
    status = "active" if (user.is_active and (not profile or profile.is_active)) else "deactivated"
    if profile and not getattr(profile, "email_verified", True):
        status = "invited"
    return {
        "id": str(user.id),
        "first_name": user.first_name or "",
        "last_name": user.last_name or "",
        "email": user.email or "",
        "username": user.username or "",
        "title": title,
        "role_id": role,
        "status": status,
        "mfa_enabled": getattr(profile, "two_factor_enabled", False) if profile else False,
        "api_access": (profile.user_settings.get("api_access", False) if profile and isinstance(profile.user_settings, dict) else False),
    }


@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated, HasFeaturePermission("account.overview.view")])
def org_user_list(request):
    """List (GET) or create (POST) users in current org."""
    from django.contrib.auth.models import User
    from users.models import UserProfile

    org_id = _get_organization_id(request)
    if not org_id:
        return Response({"error": "No organization assigned."}, status=status.HTTP_404_NOT_FOUND)

    if request.method == "GET":
        users = (
            User.objects.using("core")
            .filter(profile__organization_id=org_id)
            .select_related("profile")
            .order_by("last_name", "first_name")
        )
        return Response([_user_to_org_json(u) for u in users])

    if request.method == "POST":
        email = (request.data.get("email") or "").strip()
        if not email:
            return Response({"error": "email is required."}, status=status.HTTP_400_BAD_REQUEST)
        if User.objects.filter(email__iexact=email).exists():
            return Response({"error": "A user with this email already exists."}, status=status.HTTP_400_BAD_REQUEST)
        username = (request.data.get("username") or "").strip() or email.split("@")[0]
        if User.objects.filter(username__iexact=username).exists():
            username = f"{username}_{User.objects.count()}"
        first_name = (request.data.get("first_name") or "").strip()
        last_name = (request.data.get("last_name") or "").strip()
        role = (request.data.get("role_id") or request.data.get("role") or "viewer").lower()
        title = (request.data.get("title") or "").strip()
        password = request.data.get("password") or get_random_string(length=12)
        org_uuid = uuid_module.UUID(str(org_id))
        user = User.objects.create_user(
            username=username,
            email=email,
            password=password,
            first_name=first_name,
            last_name=last_name,
        )
        profile = UserProfile.objects.using("core").create(
            user=user,
            role=role,
            is_active=True,
            organization_id=org_uuid,
            user_settings={"title": title, "api_access": bool(request.data.get("api_access", False))},
        )
        return Response(_user_to_org_json(user), status=status.HTTP_201_CREATED)


@api_view(["GET", "PATCH", "DELETE"])
@permission_classes([IsAuthenticated, HasFeaturePermission("account.overview.view")])
def org_user_detail(request, user_id):
    """Get, update, or remove user from org."""
    from django.contrib.auth.models import User
    from users.models import UserProfile

    org_id = _get_organization_id(request)
    if not org_id:
        return Response({"error": "No organization assigned."}, status=status.HTTP_404_NOT_FOUND)

    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return Response({"error": "User not found."}, status=status.HTTP_404_NOT_FOUND)

    profile = getattr(user, "profile", None)
    if not profile or str(profile.organization_id) != str(org_id):
        return Response({"error": "User is not in this organization."}, status=status.HTTP_404_NOT_FOUND)

    if request.method == "GET":
        return Response(_user_to_org_json(user))

    if request.method == "PATCH":
        if "first_name" in request.data:
            user.first_name = (request.data.get("first_name") or "").strip()
        if "last_name" in request.data:
            user.last_name = (request.data.get("last_name") or "").strip()
        if "email" in request.data:
            email = (request.data.get("email") or "").strip()
            if email and email != user.email:
                if User.objects.filter(email__iexact=email).exclude(id=user.id).exists():
                    return Response({"error": "Email already in use."}, status=status.HTTP_400_BAD_REQUEST)
                user.email = email
        if "username" in request.data:
            user.username = (request.data.get("username") or "").strip() or user.username
        if "title" in request.data:
            settings = dict(profile.user_settings or {})
            settings["title"] = (request.data.get("title") or "").strip()
            profile.user_settings = settings
        if "role_id" in request.data or "role" in request.data:
            role = (request.data.get("role_id") or request.data.get("role") or profile.role).lower()
            from users.role_sync import set_user_role_by_name
            set_user_role_by_name(user, role)
            profile.role = role
        if "status" in request.data:
            status_val = (request.data.get("status") or "").lower()
            profile.is_active = status_val in ("active", "invited")
            user.is_active = profile.is_active
        if "api_access" in request.data:
            settings = dict(profile.user_settings or {})
            settings["api_access"] = bool(request.data.get("api_access"))
            profile.user_settings = settings
        user.save()
        profile.save()
        return Response(_user_to_org_json(user))

    if request.method == "DELETE":
        profile.organization_id = None
        profile.save()
        return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(["GET"])
@permission_classes([IsAuthenticated, HasFeaturePermission("account.overview.view")])
def org_user_bulk_template(request):
    """Download a sample Excel template for bulk user upload."""
    from io import BytesIO
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    headers = ["email", "first_name", "last_name", "username", "title", "role"]
    ws.append(headers)
    ws.append(["jane.doe@example.com", "Jane", "Doe", "jane", "Analyst", "analyst"])
    ws.append(["john.smith@example.com", "John", "Smith", "", "Manager", "manager"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 12

    # Instructions sheet
    inst = wb.create_sheet("Instructions", 0)
    inst.append(["Bulk User Import - Instructions"])
    inst["A1"].font = Font(bold=True, size=12)
    inst.append([])
    inst.append(["1. Fill in the 'Users' sheet. Email is required; other columns are optional."])
    inst.append(["2. role: viewer | analyst | manager | director | executive | admin (default: viewer)"])
    inst.append(["3. username: leave blank to auto-generate from email"])
    inst.append(["4. Save as .xlsx and upload via Users & Access > Upload users"])
    inst.append([])
    inst.append(["Columns:", "Description"])
    inst.append(["email", "Required. Must be unique."])
    inst.append(["first_name", "Optional"])
    inst.append(["last_name", "Optional"])
    inst.append(["username", "Optional. Auto-generated from email if blank."])
    inst.append(["title", "Optional job title"])
    inst.append(["role", "viewer, analyst, manager, director, executive, or admin"])
    inst.column_dimensions["A"].width = 50
    inst.column_dimensions["B"].width = 40

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="users_bulk_template.xlsx"'
    return response


@api_view(["GET"])
@permission_classes([IsAuthenticated, HasFeaturePermission("account.overview.view")])
def org_user_bulk_template_csv(request):
    """Download a CSV template for bulk user upload. Headers match auth_user + users_userprofile."""
    from django.http import HttpResponse
    import csv
    from io import StringIO

    buffer = StringIO()
    writer = csv.writer(buffer)
    headers = ["email", "first_name", "last_name", "username", "title", "role"]
    writer.writerow(headers)
    writer.writerow(["jane.doe@example.com", "Jane", "Doe", "jane", "Analyst", "analyst"])
    writer.writerow(["john.smith@example.com", "John", "Smith", "", "Manager", "manager"])
    response = HttpResponse(buffer.getvalue(), content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="users_bulk_template.csv"'
    return response


def _parse_bulk_upload_file(file_obj):
    """Parse Excel or CSV file into list of rows. Returns (rows, error_msg)."""
    import csv
    from openpyxl import load_workbook

    name = (file_obj.name or "").lower()
    if name.endswith(".csv"):
        try:
            content = file_obj.read().decode("utf-8-sig")
            reader = csv.reader(content.splitlines())
            rows = list(reader)
        except Exception as e:
            return None, f"Invalid CSV: {str(e)}"
    elif name.endswith((".xlsx", ".xls")):
        try:
            data = file_obj.read()
            from io import BytesIO
            wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
            users_sheet = next((s for s in wb.sheetnames if s.lower() == "users"), None)
            ws = wb[users_sheet] if users_sheet else wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception as e:
            return None, f"Invalid Excel: {str(e)}"
    else:
        return None, "File must be .csv, .xlsx, or .xls."
    return rows, None


@api_view(["POST"])
@permission_classes([IsAuthenticated, HasFeaturePermission("account.overview.view")])
def org_user_bulk_upload(request):
    """Upload a CSV or Excel file to bulk-create users in the current org."""
    import logging
    logger = logging.getLogger(__name__)
    try:
        return _org_user_bulk_upload_impl(request)
    except Exception as e:
        logger.exception("Bulk upload failed")
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def _ensure_sequence(conn, table_name, schema="core"):
    """Reset id sequence for a table if out of sync (fixes 'null value in column id' on INSERT)."""
    try:
        qualified = f"{schema}.{table_name}" if schema else table_name
        with conn.cursor() as c:
            c.execute("SELECT pg_get_serial_sequence(%s, 'id')", [qualified])
            row = c.fetchone()
            seq = row[0] if row else None
            seq_name = f"{schema}.{table_name}_id_seq" if schema else f"{table_name}_id_seq"
            if not seq:
                c.execute(
                    f"CREATE SEQUENCE IF NOT EXISTS {seq_name} OWNED BY {qualified}.id"
                )
                c.execute(
                    f"ALTER TABLE {qualified} "
                    f"ALTER COLUMN id SET DEFAULT nextval(%s::regclass)",
                    [seq_name],
                )
                seq = seq_name
            c.execute(f"SELECT COALESCE(MAX(id), 0) FROM {qualified}")
            max_id = c.fetchone()[0]
            c.execute("SELECT setval(%s, %s, true)", [seq, max_id])
    except Exception:
        pass


def _org_user_bulk_upload_impl(request):
    from django.contrib.auth.models import User
    from django.db import connections
    from users.models import UserProfile

    conn = connections["core"]
    _ensure_sequence(conn, "auth_user")
    _ensure_sequence(conn, "users_userprofile")
    _ensure_sequence(conn, "auth_user_groups")
    org_id = _get_organization_id(request)
    if not org_id:
        return Response({"error": "No organization found."}, status=status.HTTP_404_NOT_FOUND)
    org_uuid = uuid_module.UUID(str(org_id))

    file_obj = request.FILES.get("file")
    if not file_obj:
        return Response({"error": "No file uploaded. Use form field 'file'."}, status=status.HTTP_400_BAD_REQUEST)

    rows, parse_err = _parse_bulk_upload_file(file_obj)
    if parse_err:
        return Response({"error": parse_err}, status=status.HTTP_400_BAD_REQUEST)
    if not rows:
        return Response({"error": "File is empty."}, status=status.HTTP_400_BAD_REQUEST)

    header_row = [str(c).strip().lower() if c else "" for c in rows[0]]
    col_map = {}
    for i, h in enumerate(header_row):
        if h in ("email", "first_name", "last_name", "username", "title", "role"):
            col_map[h] = i

    if "email" not in col_map:
        return Response({"error": "Sheet must have an 'email' column."}, status=status.HTTP_400_BAD_REQUEST)

    created = []
    errors = []
    skipped = []
    for idx, row in enumerate(rows[1:], start=2):
        if not row or all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            skipped.append({"row": idx, "reason": "Empty row"})
            continue
        email = (row[col_map.get("email", 0)] or "").strip()
        if not email:
            skipped.append({"row": idx, "reason": "No email"})
            continue
        if "@" not in email:
            skipped.append({"row": idx, "email": email, "reason": "Invalid email (no @)"})
            continue
        first_name = (row[col_map.get("first_name", -1)] or "").strip() if "first_name" in col_map else ""
        last_name = (row[col_map.get("last_name", -1)] or "").strip() if "last_name" in col_map else ""
        username = (row[col_map.get("username", -1)] or "").strip() if "username" in col_map else ""
        title = (row[col_map.get("title", -1)] or "").strip() if "title" in col_map else ""
        role = (row[col_map.get("role", -1)] or "viewer").strip().lower() if "role" in col_map else "viewer"

        if User.objects.filter(email__iexact=email).exists():
            errors.append({"row": idx, "email": email, "message": "Email already exists"})
            continue

        username = username or email.split("@")[0]
        if User.objects.filter(username__iexact=username).exists():
            username = f"{username}_{User.objects.count()}"

        user = None
        try:
            from users.role_sync import set_user_role_by_name
            password = get_random_string(length=12)
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name,
            )
            UserProfile.objects.using("core").create(
                user=user,
                role=role,
                is_active=True,
                organization_id=org_uuid,
                user_settings={"title": title, "api_access": False},
            )
            set_user_role_by_name(user, role)
            created.append({"email": email, "username": user.username})
        except Exception as e:
            if user:
                try:
                    user.delete()
                except Exception:
                    pass
            errors.append({"row": idx, "email": email, "message": str(e)})

    return Response({
        "created": len(created),
        "errors": len(errors),
        "skipped": len(skipped),
        "users": created,
        "details": errors,
        "skipped_details": skipped,
    }, status=status.HTTP_200_OK)


# --- Org-scoped departments (Users & Access) ---

@api_view(["GET"])
@permission_classes([IsAuthenticated, HasFeaturePermission("account.overview.view")])
def org_department_list(request):
    """List departments for current org. Used for team type dropdown."""
    from platform_org.models import Department

    org_id = _get_organization_id(request)
    if not org_id:
        return Response({"error": "No organization assigned."}, status=status.HTTP_404_NOT_FOUND)

    depts = (
        Department.objects.using("org")
        .filter(organization_id=org_id)
        .order_by("name")
        .values("id", "department_code", "name")
    )
    return Response([
        {"id": str(d["id"]), "code": d["department_code"] or "", "name": d["name"]}
        for d in depts
    ])


# --- Org-scoped teams (Users & Access) ---

def _team_to_json(team):
    """Serialize team for API."""
    member_ids = list(
        team.members.values_list("user_id", flat=True)
    )
    dept = getattr(team, "department", None)
    return {
        "id": str(team.id),
        "name": team.name,
        "description": team.description or "",
        "team_type": team.team_type,
        "department_id": str(team.department_id) if team.department_id else None,
        "department_name": dept.name if dept else None,
        "active": team.active,
        "member_ids": [str(uid) for uid in member_ids],
    }


@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated, HasFeaturePermission("account.overview.view")])
def org_team_list(request):
    """List (GET) or create (POST) teams for current org."""
    from platform_org.models import Team, TeamMember

    org_id = _get_organization_id(request)
    if not org_id:
        return Response({"error": "No organization assigned."}, status=status.HTTP_404_NOT_FOUND)

    if request.method == "GET":
        teams = (
            Team.objects.using("org")
            .filter(organization_id=org_id)
            .select_related("department")
            .prefetch_related("members")
            .order_by("name")
        )
        return Response([_team_to_json(t) for t in teams])

    if request.method == "POST":
        name = (request.data.get("name") or "").strip()
        if not name:
            return Response({"error": "name is required."}, status=status.HTTP_400_BAD_REQUEST)
        dept_id_raw = request.data.get("department_id")
        dept_id = None
        if dept_id_raw:
            try:
                from platform_org.models import Department
                dept = Department.objects.using("org").filter(
                    id=dept_id_raw, organization_id=org_id
                ).first()
                dept_id = dept.id if dept else None
            except (ValueError, TypeError):
                pass

        team = Team.objects.using("org").create(
            organization_id=org_id,
            name=name,
            description=(request.data.get("description") or "").strip(),
            team_type=(request.data.get("team_type") or "custom").lower(),
            department_id=dept_id,
            active=bool(request.data.get("active", True)),
        )
        member_ids = request.data.get("member_ids") or []
        for uid in member_ids:
            try:
                TeamMember.objects.using("org").get_or_create(
                    team_id=team.id,
                    user_id=int(uid),
                )
            except (ValueError, TypeError):
                pass
        team = (
            Team.objects.using("org")
            .select_related("department")
            .prefetch_related("members")
            .get(id=team.id)
        )
        return Response(_team_to_json(team), status=status.HTTP_201_CREATED)


@api_view(["GET", "PATCH", "DELETE"])
@permission_classes([IsAuthenticated, HasFeaturePermission("account.overview.view")])
def org_team_detail(request, team_id):
    """Get, update, or delete a team."""
    from platform_org.models import Team, TeamMember

    org_id = _get_organization_id(request)
    if not org_id:
        return Response({"error": "No organization assigned."}, status=status.HTTP_404_NOT_FOUND)

    team = (
        Team.objects.using("org")
        .filter(id=team_id, organization_id=org_id)
        .select_related("department")
        .prefetch_related("members")
        .first()
    )
    if not team:
        return Response({"error": "Team not found."}, status=status.HTTP_404_NOT_FOUND)

    if request.method == "GET":
        return Response(_team_to_json(team))

    if request.method == "PATCH":
        if "name" in request.data:
            team.name = (request.data.get("name") or "").strip() or team.name
        if "description" in request.data:
            team.description = (request.data.get("description") or "").strip()
        if "team_type" in request.data:
            team.team_type = (request.data.get("team_type") or "custom").lower()
        if "department_id" in request.data:
            dept_id_raw = request.data.get("department_id")
            dept_id = None
            if dept_id_raw:
                try:
                    from platform_org.models import Department
                    dept = Department.objects.using("org").filter(
                        id=dept_id_raw, organization_id=org_id
                    ).first()
                    dept_id = dept.id if dept else None
                except (ValueError, TypeError):
                    pass
            team.department_id = dept_id
        if "active" in request.data:
            team.active = bool(request.data.get("active"))
        if "member_ids" in request.data:
            TeamMember.objects.using("org").filter(team_id=team.id).delete()
            for uid in (request.data.get("member_ids") or []):
                try:
                    TeamMember.objects.using("org").create(team_id=team.id, user_id=int(uid))
                except (ValueError, TypeError):
                    pass
        team.save(using="org")
        team = (
            Team.objects.using("org")
            .select_related("department")
            .prefetch_related("members")
            .get(id=team.id)
        )
        return Response(_team_to_json(team))

    if request.method == "DELETE":
        team.delete(using="org")
        return Response(status=status.HTTP_204_NO_CONTENT)


# ---------------------------------------------------------------------------
# Personnel (org.personnel) — people directory
# ---------------------------------------------------------------------------

def _personnel_to_json(p, departments=None):
    role = None
    if p.user_id:
        try:
            from users.models import UserProfile
            profile = UserProfile.objects.using("core").filter(user_id=p.user_id).first()
            role = profile.role if profile else None
        except Exception:
            pass
    return {
        "id": str(p.id),
        "first_name": p.first_name,
        "last_name": p.last_name,
        "email": p.email,
        "job_title": p.job_title or "",
        "employment_status": p.employment_status,
        "department": p.department or "",
        "external_hr_id": p.external_hr_id or "",
        "hire_date": p.hire_date.isoformat() if p.hire_date else None,
        "termination_date": p.termination_date.isoformat() if p.termination_date else None,
        "user_id": p.user_id,
        "has_access": p.user_id is not None,
        "role": role,
    }


@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated, HasFeaturePermission("account.overview.view")])
def personnel_list(request):
    """List all personnel for the org, or create a new personnel record."""
    from platform_org.models import Personnel, Department

    org_id = _get_organization_id(request)
    if not org_id:
        return Response({"error": "No organization assigned."}, status=status.HTTP_404_NOT_FOUND)

    if request.method == "GET":
        from django.contrib.auth.models import User as AuthUser
        qs = list(Personnel.objects.using("org").filter(organization_id=org_id).order_by("last_name", "first_name"))
        # Auto-reconcile: link any unlinked personnel to existing auth.User by email
        unlinked = [p for p in qs if not p.user_id and p.email]
        if unlinked:
            emails = [p.email.lower() for p in unlinked]
            matched_users = {u.email.lower(): u for u in AuthUser.objects.filter(email__in=emails)}
            for p in unlinked:
                match = matched_users.get(p.email.lower())
                if match:
                    p.user_id = match.id
                    p.save(using="org")
        return Response([_personnel_to_json(p) for p in qs])

    if request.method == "POST":
        email = (request.data.get("email") or "").strip()
        if not email:
            return Response({"error": "email is required."}, status=status.HTTP_400_BAD_REQUEST)
        if Personnel.objects.using("org").filter(organization_id=org_id, email__iexact=email).exists():
            return Response({"error": "A person with this email already exists."}, status=status.HTTP_400_BAD_REQUEST)

        import uuid as _uuid

        hire_date = None
        raw_hire = (request.data.get("hire_date") or "").strip()
        if raw_hire:
            from datetime import date
            try:
                hire_date = date.fromisoformat(raw_hire)
            except ValueError:
                pass

        p = Personnel.objects.using("org").create(
            organization_id=org_id,
            first_name=(request.data.get("first_name") or "").strip(),
            last_name=(request.data.get("last_name") or "").strip(),
            email=email,
            job_title=(request.data.get("job_title") or "").strip(),
            employment_status=(request.data.get("employment_status") or "active"),
            external_hr_id=(request.data.get("external_hr_id") or "").strip(),
            department=(request.data.get("department") or "").strip(),
            hire_date=hire_date,
        )
        return Response(_personnel_to_json(p), status=status.HTTP_201_CREATED)


@api_view(["GET", "PATCH", "DELETE"])
@permission_classes([IsAuthenticated, HasFeaturePermission("account.overview.view")])
def personnel_detail(request, personnel_id):
    """Get, update, or delete a personnel record."""
    from platform_org.models import Personnel

    org_id = _get_organization_id(request)
    if not org_id:
        return Response({"error": "No organization assigned."}, status=status.HTTP_404_NOT_FOUND)

    try:
        p = Personnel.objects.using("org").get(id=personnel_id, organization_id=org_id)
    except Personnel.DoesNotExist:
        return Response({"error": "Person not found."}, status=status.HTTP_404_NOT_FOUND)

    if request.method == "GET":
        return Response(_personnel_to_json(p))

    if request.method == "PATCH":
        if "first_name" in request.data:
            p.first_name = (request.data.get("first_name") or "").strip()
        if "last_name" in request.data:
            p.last_name = (request.data.get("last_name") or "").strip()
        if "email" in request.data:
            email = (request.data.get("email") or "").strip()
            if email and email.lower() != p.email.lower():
                if Personnel.objects.using("org").filter(organization_id=org_id, email__iexact=email).exclude(id=p.id).exists():
                    return Response({"error": "Email already in use."}, status=status.HTTP_400_BAD_REQUEST)
                p.email = email
        if "job_title" in request.data:
            p.job_title = (request.data.get("job_title") or "").strip()
        if "employment_status" in request.data:
            p.employment_status = (request.data.get("employment_status") or "active")
        if "external_hr_id" in request.data:
            p.external_hr_id = (request.data.get("external_hr_id") or "").strip()
        if "department" in request.data:
            p.department = (request.data.get("department") or "").strip()
        if "hire_date" in request.data:
            raw_hire = (request.data.get("hire_date") or "").strip()
            from datetime import date
            p.hire_date = date.fromisoformat(raw_hire) if raw_hire else None
        p.save(using="org")
        # Also update role on UserProfile if provided and person has access
        if "role" in request.data and p.user_id:
            try:
                from users.models import UserProfile
                from users.role_sync import set_user_role_by_name
                from django.contrib.auth.models import User
                role = (request.data.get("role") or "").strip().lower()
                if role:
                    auth_user = User.objects.get(id=p.user_id)
                    profile = getattr(auth_user, "profile", None)
                    if profile:
                        profile.role = role
                        profile.save()
                    set_user_role_by_name(auth_user, role)
            except Exception:
                pass
        return Response(_personnel_to_json(p))

    if request.method == "DELETE":
        if p.user_id:
            return Response({"error": "Cannot delete a person who has system access. Revoke access first."}, status=status.HTTP_400_BAD_REQUEST)
        p.delete(using="org")
        return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(["POST"])
@permission_classes([IsAuthenticated, HasFeaturePermission("account.overview.view")])
def personnel_grant_access(request, personnel_id):
    """
    Grant system access to a personnel record.
    Creates an auth.User, links it to the Personnel record via user_id.
    Payload: { role: str }
    """
    from platform_org.models import Personnel
    from django.contrib.auth.models import User
    from users.models import UserProfile
    from users.role_sync import set_user_role_by_name
    import uuid as _uuid

    org_id = _get_organization_id(request)
    if not org_id:
        return Response({"error": "No organization assigned."}, status=status.HTTP_404_NOT_FOUND)

    try:
        p = Personnel.objects.using("org").get(id=personnel_id, organization_id=org_id)
    except Personnel.DoesNotExist:
        return Response({"error": "Person not found."}, status=status.HTTP_404_NOT_FOUND)

    if p.user_id:
        return Response({"error": "This person already has system access."}, status=status.HTTP_400_BAD_REQUEST)

    role = (request.data.get("role") or "viewer").strip().lower()

    # If a user with this email already exists, link them rather than creating a duplicate
    existing_user = User.objects.filter(email__iexact=p.email).first()
    if existing_user:
        p.user_id = existing_user.id
        p.save(using="org")
        # Update role if provided
        profile = getattr(existing_user, "profile", None)
        if profile and role:
            profile.role = role
            profile.save()
            set_user_role_by_name(existing_user, role)
        return Response({
            "message": f"Linked to existing account. {p.first_name} {p.last_name} already has login access.",
            "user_id": existing_user.id,
            "username": existing_user.username,
            "role": role,
        }, status=status.HTTP_200_OK)

    username = p.email.split("@")[0]
    if User.objects.filter(username__iexact=username).exists():
        username = f"{username}_{User.objects.count()}"

    password = get_random_string(length=12)
    user = User.objects.create_user(
        username=username,
        email=p.email,
        password=password,
        first_name=p.first_name,
        last_name=p.last_name,
    )
    org_uuid = _uuid.UUID(str(org_id))
    UserProfile.objects.using("core").create(
        user=user,
        role=role,
        is_active=True,
        organization_id=org_uuid,
        user_settings={"title": p.job_title or "", "api_access": False},
    )
    set_user_role_by_name(user, role)

    # Link back to Personnel
    p.user_id = user.id
    p.save(using="org")

    return Response({
        "message": f"Access granted. {p.first_name} {p.last_name} can now log in.",
        "user_id": user.id,
        "username": user.username,
        "role": role,
        "temporary_password": password,
    }, status=status.HTTP_201_CREATED)


@api_view(["POST"])
@permission_classes([IsAuthenticated, HasFeaturePermission("account.overview.view")])
def personnel_revoke_access(request, personnel_id):
    """Revoke system access: deactivates auth.User, clears Personnel.user_id."""
    from platform_org.models import Personnel
    from django.contrib.auth.models import User

    org_id = _get_organization_id(request)
    if not org_id:
        return Response({"error": "No organization assigned."}, status=status.HTTP_404_NOT_FOUND)

    try:
        p = Personnel.objects.using("org").get(id=personnel_id, organization_id=org_id)
    except Personnel.DoesNotExist:
        return Response({"error": "Person not found."}, status=status.HTTP_404_NOT_FOUND)

    if not p.user_id:
        return Response({"error": "This person does not have system access."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        auth_user = User.objects.get(id=int(p.user_id))
        auth_user.is_active = False
        auth_user.save()
        profile = getattr(auth_user, "profile", None)
        if profile:
            profile.is_active = False
            profile.save()
    except (User.DoesNotExist, ValueError, TypeError):
        pass

    p.user_id = None
    p.save(using="org")
    return Response({"message": "Access revoked."})


@api_view(["POST"])
@permission_classes([IsAuthenticated, HasFeaturePermission("account.overview.view")])
def personnel_csv_import(request):
    """
    Import people from CSV/XLSX into org.personnel.
    Columns: first_name, last_name, email, job_title, department, employment_status, hire_date, external_hr_id
    Does NOT create auth.User — access must be granted separately.
    """
    from platform_org.models import Personnel, Department
    import uuid as _uuid

    org_id = _get_organization_id(request)
    if not org_id:
        return Response({"error": "No organization assigned."}, status=status.HTTP_404_NOT_FOUND)

    file_obj = request.FILES.get("file")
    if not file_obj:
        return Response({"error": "No file uploaded. Use form field 'file'."}, status=status.HTTP_400_BAD_REQUEST)

    rows, parse_err = _parse_bulk_upload_file(file_obj)
    if parse_err:
        return Response({"error": parse_err}, status=status.HTTP_400_BAD_REQUEST)
    if not rows:
        return Response({"error": "File is empty."}, status=status.HTTP_400_BAD_REQUEST)

    header_row = [str(c).strip().lower() if c else "" for c in rows[0]]
    col_map = {h: i for i, h in enumerate(header_row) if h}

    if "email" not in col_map:
        return Response({"error": "Sheet must have an 'email' column."}, status=status.HTTP_400_BAD_REQUEST)

    created, updated, errors, skipped = [], [], [], []

    for idx, row in enumerate(rows[1:], start=2):
        if not row or all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            skipped.append({"row": idx, "reason": "Empty row"})
            continue

        def _get(col):
            i = col_map.get(col)
            if i is None:
                return ""
            v = row[i]
            return str(v).strip() if v is not None else ""

        email = _get("email")
        if not email or "@" not in email:
            skipped.append({"row": idx, "reason": "Missing or invalid email"})
            continue

        first_name = _get("first_name")
        last_name = _get("last_name")
        job_title = _get("job_title")
        employment_status = _get("employment_status") or "active"
        external_hr_id = _get("external_hr_id")
        hire_date_raw = _get("hire_date")
        department = _get("department")

        hire_date = None
        if hire_date_raw:
            from datetime import date
            try:
                hire_date = date.fromisoformat(hire_date_raw)
            except ValueError:
                pass

        try:
            existing = Personnel.objects.using("org").filter(organization_id=org_id, email__iexact=email).first()
            if existing:
                existing.first_name = first_name or existing.first_name
                existing.last_name = last_name or existing.last_name
                existing.job_title = job_title or existing.job_title
                existing.employment_status = employment_status
                if external_hr_id:
                    existing.external_hr_id = external_hr_id
                if department:
                    existing.department = department
                if hire_date:
                    existing.hire_date = hire_date
                existing.save(using="org")
                updated.append({"email": email})
            else:
                Personnel.objects.using("org").create(
                    organization_id=org_id,
                    first_name=first_name,
                    last_name=last_name,
                    email=email,
                    job_title=job_title,
                    employment_status=employment_status,
                    external_hr_id=external_hr_id,
                    department=department,
                    hire_date=hire_date,
                )
                created.append({"email": email})
        except Exception as e:
            errors.append({"row": idx, "email": email, "message": str(e)})

    return Response({
        "created": len(created),
        "updated": len(updated),
        "errors": len(errors),
        "skipped": len(skipped),
        "details": errors,
    }, status=status.HTTP_200_OK)


@api_view(["GET"])
@permission_classes([IsAuthenticated, HasFeaturePermission("account.overview.view")])
def personnel_csv_template(request):
    """Download a CSV template for personnel import."""
    from django.http import HttpResponse
    import csv
    import io

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["first_name", "last_name", "email", "job_title", "department", "employment_status", "hire_date", "external_hr_id"])
    writer.writerow(["Jane", "Smith", "jane.smith@company.com", "Security Engineer", "Engineering", "active", "2023-01-15", "EMP001"])
    content = output.getvalue()
    response = HttpResponse(content, content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="people_import_template.csv"'
    return response
